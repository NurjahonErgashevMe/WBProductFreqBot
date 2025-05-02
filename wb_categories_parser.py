import json
import logging
from typing import Any, Dict, List, Optional

import pandas as pd
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook import Workbook

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("wb_parser.log"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)


class WBCategoriesParser:
    """Parser for Wildberries categories with Evirma API integration."""

    WB_CATEGORIES_URL = "https://static-basket-01.wbcontent.net/vol0/data/main-menu-uz-ru-v3.json"
    EVIRMA_API_URL = "https://evirma.ru/api/v1/keyword/list"
    DEFAULT_HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    EXCEL_FILENAME = "wb_categories_analysis.xlsx"
    JSON_FILENAME = "categories.json"

    def __init__(self, headers: Optional[Dict[str, str]] = None):
        """Initialize the parser with optional custom headers."""
        self.headers = headers or self.DEFAULT_HEADERS
        self.session = requests.Session()
        self.session.headers.update(self.headers)

    def fetch_data(self, url: str, method: str = "GET", **kwargs) -> Any:
        """Generic method to fetch data from API with error handling."""
        try:
            logger.info(f"Fetching data from {url}")
            response = self.session.request(method, url, **kwargs)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"Request failed for {url}: {e}")
            raise
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON response from {url}: {e}")
            raise

    def get_wb_categories(self) -> List[Dict[str, Any]]:
        """Fetch categories from Wildberries API."""
        return self.fetch_data(self.WB_CATEGORIES_URL)

    def extract_category_hierarchy(self, categories: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Extract category hierarchy with SEO keywords."""
        result = []

        def _process_category(
            category: Dict[str, Any],
            parent_name: Optional[str] = None,
            root_name: Optional[str] = None,
        ) -> None:
            """Recursively process category and its children."""
            current_name = category["name"]
            current_seo = category.get("seo", "")
            current_root = root_name or current_name

            result.append({"SEO": current_seo})

            for child in category.get("childs", []):
                _process_category(child, current_name, current_root)

        for category in categories:
            _process_category(category)

        return result

    def get_evirma_data(self, keywords: List[str]) -> Dict[str, Any]:
        """Fetch keyword statistics from Evirma API."""
        payload = {"keywords": keywords, "an": False}
        data = self.fetch_data(self.EVIRMA_API_URL, method="POST", json=payload, timeout=30)

        # Filter keywords with product_count > 0
        filtered_keywords = {
            keyword: stats
            for keyword, stats in data.get("data", {}).get("keywords", {}).items()
            if stats
            and isinstance(stats, dict)
            and (stats.get("product_count", 0) or 0) > 0
        }
        data["data"]["keywords"] = filtered_keywords

        self._save_json(data, self.JSON_FILENAME)
        return data

    def _save_json(self, data: Dict[str, Any], filename: str) -> None:
        """Save data to JSON file."""
        try:
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
            logger.info(f"Data saved to {filename}")
        except IOError as e:
            logger.error(f"Failed to save JSON to {filename}: {e}")
            raise

    def merge_data(
        self, categories: List[Dict[str, Any]], evirma_data: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Merge Wildberries categories with Evirma statistics."""
        merged = []
        keywords_data = evirma_data.get("data", {}).get("keywords", {})

        for keyword, stats in keywords_data.items():
            if not isinstance(stats, dict):
                continue

            # Extract product count from different possible locations
            product_count = (
                stats.get("product_count", 0)
                or stats.get("cluster", {}).get("product_count", 0)
            )

            # Extract frequency data
            freq_data = stats.get("freq", {})
            cluster_freq = stats.get("cluster", {}).get("freq_common", {})
            
            merged.append({
                "Keyword": keyword,
                "Product Count": product_count,
                "Yearly Frequency": stats.get("freq365", 0) or cluster_freq.get("keyword_count", 0),
                "Monthly Frequency": freq_data.get("monthly", 0) or cluster_freq.get("monthly", 0),
                "Weekly Frequency": freq_data.get("weekly", 0) or cluster_freq.get("weekly", 0),
                "Weekly Trend": freq_data.get("weekly_trend", 0) or cluster_freq.get("weekly_trend", 0),
            })

        return merged

    def save_to_excel(self, data: List[Dict[str, Any]], filename: Optional[str] = None) -> None:
        """Save data to Excel with formatting."""
        filename = filename or self.EXCEL_FILENAME
        df = pd.DataFrame(data)
        df = df.sort_values(by=["Monthly Frequency"], ascending=False)

        try:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Category Analysis")
                self._format_excel(writer, df)
            logger.info(f"Excel report saved to {filename}")
        except Exception as e:
            logger.error(f"Failed to save Excel file {filename}: {e}")
            raise

    def _format_excel(self, writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
        """Apply formatting to Excel worksheet."""
        workbook = writer.book
        worksheet = writer.sheets["Category Analysis"]

        # Column widths
        column_widths = {
            "A": 50,  # Keyword
            "B": 15,  # Product Count
            "C": 20,  # Yearly Frequency
            "D": 20,  # Weekly Frequency
            "E": 20,  # Monthly Frequency
            "F": 15,  # Weekly Trend
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Header formatting
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")

        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = header_alignment

        # Color coding
        colors = {
            "high": PatternFill(start_color="90EE90", fill_type="solid"),  # Light green
            "medium": PatternFill(start_color="FFFFE0", fill_type="solid"),  # Light yellow
            "low": PatternFill(start_color="FFB6C1", fill_type="solid"),    # Light red
        }

        total_rows = len(df)
        high_cutoff = total_rows // 3
        medium_cutoff = 2 * (total_rows // 3)

        for row in range(2, worksheet.max_row + 1):
            if row <= high_cutoff + 1:
                fill = colors["high"]
            elif row <= medium_cutoff + 1:
                fill = colors["medium"]
            else:
                fill = colors["low"]

            for cell in worksheet[row]:
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

    def run(self) -> bool:
        """Execute the full parsing pipeline."""
        try:
            logger.info("Starting Wildberries categories parsing")
            
            # 1. Fetch and process categories
            categories = self.get_wb_categories()
            category_hierarchy = self.extract_category_hierarchy(categories)
            logger.info(f"Processed {len(category_hierarchy)} categories")

            # 2. Get SEO keywords and fetch Evirma data
            seo_keywords = [cat["SEO"] for cat in category_hierarchy if cat["SEO"]]
            evirma_data = self.get_evirma_data(seo_keywords)

            # 3. Merge and save results
            merged_data = self.merge_data(category_hierarchy, evirma_data)
            self.save_to_excel(merged_data)

            logger.info("Processing completed successfully")
            return True
        except Exception as e:
            logger.error(f"Processing failed: {e}", exc_info=True)
            return False


if __name__ == "__main__":
    parser = WBCategoriesParser()
    success = parser.run()
    exit(0 if success else 1)